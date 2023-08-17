from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from datetime import datetime
import openpyxl
import sys
import os

class excel_automation:

    def __init__(self, filename):
        self.filename = filename


    def open_excel(self, file_path, sheet_num=0):
        """
        Starts an excel session
        """
        try:

            if not os.path.exists(file_path):
                wb = openpyxl.Workbook()
                wb.save(file_path)
            else:
                wb = load_workbook(filename=file_path)
            ws = wb.worksheets[sheet_num]

            return wb, ws
        except Exception as err:
            exc_tb = sys.exc_info()[2]
            raise Exception(f'{err} ({exc_tb.tb_lineno})')
        

    def set_cell_format(self, value):
        """
        Format the cell based on the type of the value
        """
        try:
            if type(value) == float:
                return '#,##0.00'
            elif type(value) == datetime:
                return 'dd/mm/yyyy'
            else:
                return '@'
        except Exception as err:
            exc_tb = sys.exc_info()[2]
            raise Exception(f'{err} ({exc_tb.tb_lineno})')
        

    def get_next_row(self, ws, col_num: int = 1):
        """
        Returns the index of the next empty row
        """
        try:
            row_no = ws.max_row 
            loop_controller = True
            while loop_controller:

                cell_value = ws.cell(row=row_no, column=col_num).value
                if cell_value is not None:
                    if type(cell_value) != str:
                        try:
                            cell_value = str(cell_value)
                        except TypeError as err:
                            raise Exception(f'{err}, last value: {cell_value}')
                    if len(cell_value) > 0:
                        return row_no + 1
                row_no -= 1
        except Exception as err:
            exc_tb = sys.exc_info()[2]
            raise Exception(f'{err} ({exc_tb.tb_lineno})')


    def get_next_col(self, ws, row_num: int = 1):
        """
        Returns the index of the next empty column
        """
        try:
            col_no = ws.max_column
            loop_controller = True
            while loop_controller:

                cell_value = ws.cell(row=row_num, column=col_no).value

                if cell_value is not None:
                    if type(cell_value) != str:
                        try:
                            cell_value = str(cell_value)
                        except TypeError as err:
                            raise Exception(f'{err}, last value: {cell_value}')
                    if len(cell_value) > 0:
                        return col_no + 1
                col_no -= 1
        except Exception as err:
            exc_tb = sys.exc_info()[2]
            raise Exception(f'{err} ({exc_tb.tb_lineno})')
        

    def write_excel(self, wb, ws, data_array: list,
                    init_row: int = None, init_col: int = None,
                    highlight_elements: bool = False, isHeader: bool = False):
        """
        Writes an array inside an excel file
        """
        try:
            content_font = Font(name='Aptos', size=11, bold=False,
                                italic=False, vertAlign=None, underline='none',
                                strike=False, color='000000')

            content_alignment = Alignment(horizontal='center', vertical='top',
                                        text_rotation=0, wrap_text=False,
                                        shrink_to_fit=False, indent=0)

            if highlight_elements:
                highlight_fill = PatternFill(fill_type='solid',
                                            start_color='FABF8F',
                                            end_color='FABF8F')

            if isHeader:
                header_font = Font(name='Aptos', size=11, bold=True,
                                italic=False, vertAlign=None, underline='none',
                                strike=False, color='000000')

            if init_row is None:
                init_row = self.get_next_row(ws)
            if init_row < 0:
                init_row_aux = self.get_next_row(ws)
                init_row = init_row_aux + init_row

            if init_col is None:
                init_col = self.get_next_col(ws)
            if init_col < 0:
                init_col_aux = self.get_next_col(ws)
                init_col = init_col_aux + init_col

            if data_array:
                for i, row in enumerate(data_array, start=init_row):
                    # Cell Height
                    # ws.row_dimensions[i].height = 26
                    for j, cell_value in enumerate(row, start=init_col):
                        if isHeader:
                            # Cell Width
                            ws.column_dimensions[get_column_letter(j)].width = len(str(cell_value)) + 3

                        if type(cell_value) == str:
                            cell_value = cell_value.strip()

                        #check hyperlinks
                        if type(cell_value) == str and (":\\" in cell_value or "/home" in cell_value):
                            # cell = ws.cell(column=j, row=i, value='=HYPERLINK("{}", "{}")'.format(cell_value, cell_value.split("/")[-1])).style = 'Hyperlink'
                            cell = ws.cell(column=j, row=i, value=cell_value.split("/")[-1])

                        else:
                            cell = ws.cell(column=j, row=i,
                                        value=cell_value)

                            cell.number_format = self.set_cell_format(cell_value)

                            if highlight_elements:
                                cell.fill = highlight_fill

                            if isHeader:
                                cell.font = header_font
                            else:
                                cell.font = content_font

                            cell.alignment = content_alignment
            else:
                # Cell Height
                # ws.row_dimensions[init_row].height = 26
                cell = ws.cell(column=init_col, row=init_row)
                cell.number_format = self.set_cell_format('str')
                cell.value = 'No information found for this section'
                cell.font = content_font
                cell.alignment = content_alignment
        except Exception as err:
            exc_tb = sys.exc_info()[2]
            raise Exception(f'{err} ({exc_tb.tb_lineno})')


